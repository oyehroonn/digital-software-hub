#!/usr/bin/env python
"""Quick script to check Excel file structure"""
import sys

try:
    from openpyxl import load_workbook
    wb = load_workbook("products.xlsx", data_only=True)
    ws = wb.active
    
    print("First row (headers):")
    headers = [cell.value for cell in ws[1]]
    print(headers)
    
    print("\nFirst few data rows:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
        print(f"Row {i+1}: {row}")
    
    print(f"\nTotal rows: {ws.max_row}")
except ImportError:
    print("openpyxl not installed. Please install it:")
    print("  pip install --user openpyxl")
    sys.exit(1)
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()

