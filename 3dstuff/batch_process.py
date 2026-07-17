#!/usr/bin/env python
"""
Batch processor: loops through ALL products in products.xlsx,
creates folders in models/, extracts the EMBEDDED image that is
anchored to each product's row, applies it as a texture on box.glb.

Emits BOTH:
  1. models/{id}_{short_name}/model.glb   (folder layout, consumed by api.py)
  2. public/models/{id}.glb               (FLAT layout, what the website loads)
  3. models/manifest.json                 (id → name → sku → file, + api fields)

IMPORTANT (image mapping fix):
  Images are matched to products by their DRAWING ANCHOR (the row the picture
  actually sits on) — NOT by the alphabetical order of the hash-named files in
  xl/media/. The media files are named by content hash, so their sorted order
  does NOT follow row order; the old index-based mapping mis-assigned textures.

This script is idempotent: products that already have a flat {id}.glb are
skipped, so it can be re-run cheaply. A full 478-model regen is a Wave-3 task —
running this now only fills gaps.
"""

import os
import sys
import io
import json
import re
import shutil
import traceback
from pathlib import Path

# ── Imports ──────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: pip install --user openpyxl")

try:
    import trimesh
except ImportError:
    sys.exit("ERROR: pip install --user trimesh")

try:
    from PIL import Image, ImageDraw, ImageFont, ImageFilter
except ImportError:
    sys.exit("ERROR: pip install --user pillow")

import html as _html

# AVIF support — some images in the xlsx are AVIF with a .jpg extension
try:
    import pillow_avif  # noqa: F401 — registers AVIF opener with Pillow
except ImportError:
    print("WARNING: pillow-avif-plugin not installed. AVIF images may fail.")
    print("  Install with: pip install --user pillow-avif-plugin")

try:
    import numpy as np
except ImportError:
    sys.exit("ERROR: pip install --user numpy")

try:
    from scipy import ndimage
except ImportError:
    sys.exit("ERROR: pip install --user scipy")


# ── Config (all paths anchored to this file's directory) ─────────────────
HERE          = Path(__file__).resolve().parent
EXCEL_PATH    = HERE / "products.xlsx"
BASE_GLB      = HERE / "box.glb"
MODELS_DIR    = HERE / "models"                       # folder layout for api.py
PUBLIC_MODELS = HERE.parent / "public" / "models"     # FLAT layout for the site
MANIFEST_PATH = MODELS_DIR / "manifest.json"
ASSETS        = HERE          # dsm.png / dsm-white.png live next to this script
MAX_FOLDER    = 60           # max chars for folder name
GLB_FILENAME  = "model.glb"  # short, predictable per-folder output name

# Fonts — DejaVu on the VPS; falls back to bundled scratchpad copies locally.
def _find_font(*cands):
    for c in cands:
        if os.path.exists(c):
            return c
    return cands[0]

FONT_BOLD = _find_font(
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    str(HERE / "fonts" / "DejaVuSans-Bold.ttf"),
)
FONT_REG = _find_font(
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    str(HERE / "fonts" / "DejaVuSans.ttf"),
)

# Cover canvas (matches gen_own_boxes.py / the DSM-Original template)
COVER_W = COVER_H = 1024
SPINE_W = 190


# ── Helpers ──────────────────────────────────────────────────────────────

def sanitize(name: str, max_len: int = MAX_FOLDER) -> str:
    """Turn a product name into a safe folder/file name."""
    name = re.sub(r'[&;]+amp;?', 'and', name)            # &amp; → and
    name = re.sub(r'[–—]', '-', name)                     # em/en dash
    name = re.sub(r'[^A-Za-z0-9 _\-]', '', name).strip()
    name = re.sub(r'\s+', '_', name)
    if len(name) > max_len:
        name = name[:max_len].rstrip('_')
    return name


# ── DSM-Original cover template (data-driven, matches gen_own_boxes.py) ────
# The catalog used to paste each product's raw WooCommerce image onto the box,
# giving wildly inconsistent quality. Instead we now DESIGN a clean, premium,
# uniform cover for every product from its data (name / category / brand): a
# white gradient face, a brand-coloured spine with the DSM logo, the DSM logo
# on the face, a fitted product name, a category/marketing tagline, a badge,
# feature bullets and a branded initials tile — the "DSM Original" look.

def hx(h):
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def clean_text(s):
    """Un-escape HTML entities and normalise dashes/whitespace for display."""
    if s is None:
        return ""
    s = _html.unescape(str(s))
    s = re.sub(r'[–—]', '-', s)      # en/em dash → hyphen
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def vgrad(size, top, bottom):
    """Vertical gradient image."""
    w, h = size
    top = np.array(top, float); bottom = np.array(bottom, float)
    t = np.linspace(0, 1, h)[:, None]
    col = (top[None, :] * (1 - t) + bottom[None, :] * t)
    arr = np.repeat(col[:, None, :], w, axis=1).astype(np.uint8)
    return Image.fromarray(arr, "RGB")


def rounded_mask(size, radius):
    m = Image.new("L", size, 0)
    ImageDraw.Draw(m).rounded_rectangle([0, 0, size[0] - 1, size[1] - 1],
                                        radius=radius, fill=255)
    return m


def fit_font(path, text, max_w, start, min_size=20):
    s = start
    while s > min_size:
        f = ImageFont.truetype(path, s)
        if f.getbbox(text)[2] <= max_w:
            return f
        s -= 2
    return ImageFont.truetype(path, min_size)


def wrap(draw, text, font, max_w):
    words, lines, cur = text.split(), [], ""
    for w_ in words:
        trial = (cur + " " + w_).strip()
        if draw.textlength(trial, font=font) <= max_w:
            cur = trial
        else:
            if cur:
                lines.append(cur)
            cur = w_
    if cur:
        lines.append(cur)
    return lines


def ellipsize(draw, text, font, max_w):
    if draw.textlength(text, font=font) <= max_w:
        return text
    while text and draw.textlength(text + "…", font=font) > max_w:
        text = text[:-1]
    return text.rstrip() + "…"


def fit_title(draw, text, max_w, start, min_size, max_lines):
    """Largest font at which `text` wraps to <= max_lines lines within max_w."""
    s = start
    while s > min_size:
        f = ImageFont.truetype(FONT_BOLD, s)
        lines = wrap(draw, text, f, max_w)
        if len(lines) <= max_lines and all(
                draw.textlength(ln, font=f) <= max_w for ln in lines):
            return f, lines
        s -= 2
    f = ImageFont.truetype(FONT_BOLD, min_size)
    lines = wrap(draw, text, f, max_w)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        last = lines[-1]
        while last and draw.textlength(last + "…", font=f) > max_w:
            last = last[:-1]
        lines[-1] = last.rstrip() + "…"
    return f, lines


# ── Brand detection ───────────────────────────────────────────────────────
# (keywords matched against name+categories, label, spine_top hex, spine_bot hex)
# Order matters: specific sub-families before their parent vendor.
BRAND_RULES = [
    (["visual studio"],                         "Visual Studio",     "4b2067", "7c3aed"),
    (["sql server"],                            "SQL Server",        "8a1220", "cc2927"),
    (["microsoft teams", "teams rooms", "teams premium"], "Microsoft Teams", "3b3c78", "6264a7"),
    (["dynamics"],                              "Microsoft Dynamics","062a5e", "0b63b8"),
    (["visio"],                                 "Microsoft Visio",   "23407a", "3955a3"),
    (["microsoft project"],                     "Microsoft Project", "1c5e1a", "31752f"),
    (["exchange"],                              "Microsoft Exchange","004c8c", "0072c6"),
    (["windows server"],                        "Windows Server",    "1e4d8c", "2f7bd0"),
    (["windows 10", "windows 11", "windows "],  "Microsoft Windows", "0050a0", "0078d4"),
    (["office 365", "microsoft 365", "office applications",
      "microsoft office", "office professional", "office standard",
      "copilot"],                               "Microsoft Office",  "a8330b", "e8590c"),
    (["microsoft"],                             "Microsoft",         "10457f", "2563eb"),
    (["autocad"],                               "AutoCAD",           "9b1b2e", "e51937"),
    (["revit"],                                 "Revit",             "045f86", "0696d7"),
    (["maya"],                                  "Maya",              "0a6f66", "00a98f"),
    (["3ds max", "3dsmax"],                     "3ds Max",           "146079", "1b98c4"),
    (["civil 3d", "civil3d"],                   "Civil 3D",          "245c34", "3a7d44"),
    (["inventor"],                              "Inventor",          "8a4b00", "d16b00"),
    (["fusion 360", "fusion360", "fusion"],     "Fusion 360",        "b34700", "ff7a00"),
    (["navisworks", "naviswork"],               "Navisworks",        "2f3f78", "4055a8"),
    (["infraworks", "infodrainage", "construction cloud",
      "aec collection", "autodesk"],            "Autodesk",          "1f2937", "4b5563"),
    (["v-ray", "vray"],                         "V-Ray",             "9c0016", "e30613"),
    (["corona"],                                "Corona Renderer",   "8a2f00", "e2571f"),
    (["enscape"],                               "Enscape",           "0a6f52", "17a67e"),
    (["chaos"],                                 "Chaos",             "8b1020", "d62839"),
    (["sketchup"],                              "SketchUp",          "9e1f17", "e5322c"),
    (["adobe", "photoshop", "illustrator", "acrobat"], "Adobe",      "8a0900", "fa0f00"),
    (["corel"],                                 "Corel",             "1f6b2e", "2f9e44"),
]
DEFAULT_BRAND = ("Digital Software Market", "1e3a8a", "2563eb")

# vendor words to strip when deriving the initials glyph
_VENDOR_WORDS = {"microsoft", "autodesk", "chaos", "adobe", "corel", "trimble"}
# generic/edition words that shouldn't drive the glyph
_STOP_WORDS = {
    "license", "licenses", "licensing", "professional", "pro", "enterprise",
    "standard", "premium", "solo", "edition", "software", "suite",
    "plus", "mak", "key", "keys", "plan", "user", "users", "multi", "volume",
    "ltsc", "for", "and", "the", "with", "made", "simple", "better", "build",
    "connected", "workflows", "real-time", "rendering", "collection", "of",
    "a", "an", "to", "your", "on", "by", "in", "at", "new",
}
# categories that are too generic to make a good subtitle
_GENERIC_CATS = {
    "uncategorized", "small business", "windows", "software",
    "agencies & freelancers software", "corporate it teams software",
    "architecture and engineer",
}


def detect_brand(name, cats):
    # Match the NAME first (the real product brand), then fall back to the
    # category — categories like "SketchUp & V-Ray" are shared ecosystem
    # buckets and must not override a name like "Chaos Corona Renderer".
    for src in (name.lower(), cats.lower()):
        for keys, label, top, bot in BRAND_RULES:
            if any(k in src for k in keys):
                return label, top, bot
    return DEFAULT_BRAND


def _split_title_tail(name):
    """Split 'Product Name : Marketing tail' / 'Name - variant' on the first
    spaced dash or colon — the head is the product title, the tail a tagline."""
    m = re.split(r'\s*:\s+|\s+-\s+', name, maxsplit=1)
    head = m[0].strip()
    tail = m[1].strip() if len(m) > 1 else ""
    return head, tail


_EDITION_RE = re.compile(
    r'^(solo|premium|pro|professional|enterprise|standard|studio|'
    r'\d+(\s*users?)?|\d+\s*/.*|[a-z]?\d+)$', re.I)


def _initials(words, n=2):
    return "".join(w[0] for w in words[:n]).upper()


def derive_glyph(head, brand_label=""):
    """A short (<=3 char) initials glyph from the significant words of a title,
    falling back to the brand name when the title is all generic tokens."""
    head = re.sub(r'\([^)]*\)', ' ', head)            # drop parentheticals
    # keep hyphenated brands whole ("V-Ray" → VR, not R3); strip edge punctuation
    words = [w.strip('.,:;®™') for w in re.split(r'[\s/&]+', head)]
    words = [w for w in words if w]
    sig = [w for w in words
           if w.lower() not in _VENDOR_WORDS
           and w.lower() not in _STOP_WORDS
           and not re.fullmatch(r'[\d.]+', w)
           and len(w) > 1]
    if not sig:
        # title was all vendor/edition tokens → use the brand name instead
        bw = [w for w in re.split(r'[\s\-/&]+', brand_label)
              if w and w.lower() not in _VENDOR_WORDS and len(w) > 1]
        if len(bw) >= 2:
            return _initials(bw, 2)
        if bw:
            w = bw[0]
            caps = [c for c in w if c.isupper()]
            return ("".join(caps[:2]) if len(caps) >= 2 else w[:2]).upper()
        bl = re.sub(r'[^A-Za-z]', '', brand_label)
        return (bl[:2] or "SW").upper()
    if len(sig) >= 2:
        return _initials(sig, 2)
    w = sig[0]
    caps = [c for c in w if c.isupper()]              # camelCase → initials
    if len(caps) >= 2:
        return "".join(caps[:2]).upper()
    return w[:2].upper()


def choose_subtitle(head, tail, cats, brand_label):
    """Prefer a descriptive marketing tail; else a meaningful category."""
    if tail and not _EDITION_RE.match(tail):
        # drop trailing variant segments ("… - Studio", "… - 5 User")
        segs = re.split(r'\s+-\s+', tail)
        while len(segs) > 1 and _EDITION_RE.match(segs[-1].strip()):
            segs.pop()
        sub = re.sub(r'\([^)]*\)', '', " - ".join(segs)).strip(' -–—')
        if sub and len(sub.split()) >= 2:
            return sub[:48]
    for c in [c.strip() for c in cats.split(',') if c.strip()]:
        if c.lower() not in _GENERIC_CATS and not re.fullmatch(r'20\d\d', c):
            # avoid subtitle that just repeats the product name
            if c.lower() not in head.lower():
                return c[:48]
    return f"{brand_label} License"


def derive_edition_badge(name):
    yrs = re.findall(r'\b(20\d\d)\b', name)
    if yrs:
        return f"{max(yrs)} EDITION"
    return "GENUINE LICENSE"


def choose_features(name, brand_label):
    low = name.lower()
    third = "Priority support"
    if any(k in low for k in ("server", "volume", "enterprise", "mak", "ltsc")):
        third = "Enterprise ready"
    elif brand_label in ("V-Ray", "Corona Renderer", "Enscape", "SketchUp",
                         "Maya", "3ds Max", "AutoCAD", "Revit", "Adobe",
                         "Fusion 360", "Civil 3D", "Inventor"):
        third = "Pro-grade tools"
    return ["Instant delivery", "Genuine license", third]


def derive_cover_params(product):
    """Turn one product row into kwargs for make_cover — never raises."""
    pid = product.get('ID')
    raw_name = product.get('Name')
    name = clean_text(raw_name) or f"Product {pid}"
    cats = clean_text(product.get('Categories'))
    brand_label, atop, abot = detect_brand(name, cats)
    head, tail = _split_title_tail(name)
    head = re.sub(r'\([^)]*\)', '', head).strip() or name
    subtitle = choose_subtitle(head, tail, cats, brand_label)
    glyph = derive_glyph(head, brand_label)
    badge = derive_edition_badge(name)
    features = choose_features(name, brand_label)
    if brand_label == "Digital Software Market":
        spine_text = "GENUINE  LICENSE"
        bottom_title = "Digital Software Market"
        bottom_sub = "Genuine · Supported · Licensed"
    else:
        spine_text = brand_label.upper()
        bottom_title = brand_label
        bottom_sub = "Genuine License · via DSM"
    return dict(name=head, tagline=subtitle, glyph=glyph, atop=atop, abot=abot,
                features=features, badge=badge, spine_text=spine_text,
                bottom_title=bottom_title, bottom_sub=bottom_sub)


def make_cover(name, tagline, glyph, atop, abot, features=None,
               badge="2026 EDITION", spine_text="A  DSM  ORIGINAL",
               bottom_title="DSM Original",
               bottom_sub="Genuine · Supported · Licensed"):
    """Render the premium DSM cover. Shared by catalog + DSM-Original boxes."""
    W = H = COVER_W
    top, bot = hx(atop), hx(abot)
    features = features or []
    img = Image.new("RGB", (W, H), (255, 255, 255))
    img.paste(vgrad((W, H), (255, 255, 255), (238, 242, 246)), (0, 0))
    d = ImageDraw.Draw(img)

    # ── brand-coloured spine (visible side of the box) ──
    img.paste(vgrad((SPINE_W, H), top, bot), (0, 0))
    try:
        dsm_w = Image.open(ASSETS / "dsm-white.png").convert("RGBA")
        r = 150 / dsm_w.width
        dsm_w = dsm_w.resize((150, int(dsm_w.height * r)), Image.LANCZOS)
        dsm_w = dsm_w.rotate(90, expand=True)
        img.paste(dsm_w, (int((SPINE_W - dsm_w.width) / 2), 70), dsm_w)
    except Exception as e:
        print("  spine logo skipped:", e)
    stf = fit_font(FONT_BOLD, spine_text, 640, 30, 18)
    strip = Image.new("RGBA", (720, 52), (0, 0, 0, 0))
    ImageDraw.Draw(strip).text((0, 0), spine_text, font=stf,
                               fill=(255, 255, 255, 235))
    strip = strip.rotate(90, expand=True)
    img.paste(strip, (int((SPINE_W - strip.width) / 2) + 4,
                      H - strip.height - 60), strip)

    mx = SPINE_W + 46

    # ── DSM horizontal logo, top-left of face ──
    try:
        dsm = Image.open(ASSETS / "dsm.png").convert("RGBA")
        r = 230 / dsm.width
        dsm = dsm.resize((230, int(dsm.height * r)), Image.LANCZOS)
        img.paste(dsm, (mx, 62), dsm)
    except Exception as e:
        print("  face logo skipped:", e)

    # ── badge pill, top-right ──
    bf = ImageFont.truetype(FONT_BOLD, 26)
    btw = d.textlength(badge, font=bf)
    bpx, bph = 26, 46
    bx1 = W - 60
    bx0 = bx1 - (btw + bpx * 2)
    by0 = 74
    d.rounded_rectangle([bx0, by0, bx1, by0 + bph], radius=bph // 2, fill=top)
    d.text((bx0 + bpx, by0 + (bph - 26) // 2 - 2), badge, font=bf,
           fill=(255, 255, 255, 245))

    # ── centre initials tile (rounded square, accent gradient, glyph) ──
    # Fixed on the right so text lives in a clean left column that never
    # collides with it — the key to a uniform, un-crowded catalogue.
    iw = 292
    ix = W - iw - 70
    iy = 432
    ic = (ix + iw // 2, iy + iw // 2)
    faint = tuple(int(c + (245 - c) * 0.82) for c in top)
    for rr in range(iw, iw + 200, 30):
        d.ellipse([ic[0] - rr // 2, ic[1] - rr // 2,
                   ic[0] + rr // 2, ic[1] + rr // 2], outline=faint, width=2)
    sh = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    ImageDraw.Draw(sh).rounded_rectangle(
        [ix + 12, iy + 20, ix + iw + 12, iy + iw + 20], radius=58,
        fill=(20, 30, 50, 70))
    sh = sh.filter(ImageFilter.GaussianBlur(22))
    img.paste(sh, (0, 0), sh)
    icon = vgrad((iw, iw), top, bot)
    icon.putalpha(rounded_mask((iw, iw), 58))
    gd = ImageDraw.Draw(icon)
    gfont = fit_font(FONT_BOLD, glyph, iw - 80, 176, 60)
    gb = gd.textbbox((0, 0), glyph, font=gfont)
    gd.text(((iw - (gb[2] - gb[0])) / 2 - gb[0],
             (iw - (gb[3] - gb[1])) / 2 - gb[1]),
            glyph, font=gfont, fill=(255, 255, 255, 245))
    img.paste(icon, (ix, iy), icon)

    # ── left column: product name (fitted, <=3 lines) ──
    col_w = ix - mx - 34          # stays clear of the tile
    name_font, lines = fit_title(d, name, col_w, 80, 32, 3)
    lh = name_font.getbbox("Ag")[3] + 14
    y = 210
    for ln in lines:
        d.text((mx, y), ln, font=name_font, fill=(24, 28, 34))
        y += lh
    title_end = y

    # accent rule + tagline
    d.rounded_rectangle([mx + 2, title_end + 6, mx + 110, title_end + 12],
                        radius=3, fill=bot)
    tag_y = title_end + 30
    if tagline:
        tag_font = fit_font(FONT_REG, tagline, col_w, 38, 24)
        d.text((mx, tag_y), ellipsize(d, tagline, tag_font, col_w),
               font=tag_font, fill=(107, 114, 128))
    fy = tag_y + (72 if tagline else 30)

    # ── feature bullets (left column) ──
    for feat in features[:3]:
        cy = fy + 15
        d.ellipse([mx, cy - 9, mx + 18, cy + 9], fill=top)
        d.line([(mx + 4, cy), (mx + 8, cy + 5), (mx + 14, cy - 5)],
               fill=(255, 255, 255), width=3, joint="curve")
        fnt = fit_font(FONT_REG, feat, col_w - 34, 28, 18)
        d.text((mx + 34, fy), ellipsize(d, feat, fnt, col_w - 34),
               font=fnt, fill=(60, 66, 74))
        fy += 52

    # ── bottom brand mark ──
    by = H - 150
    d.rounded_rectangle([mx, by, mx + 40, by + 40], radius=6, fill=top)
    d.rounded_rectangle([mx + 48, by, mx + 88, by + 40], radius=6, fill=bot)
    bt_font = fit_font(FONT_BOLD, bottom_title, W - (mx + 104) - 50, 40, 22)
    d.text((mx + 104, by - 2), bottom_title, font=bt_font, fill=(24, 28, 34))
    d.text((mx + 104, by + 46), bottom_sub,
           font=ImageFont.truetype(FONT_REG, 22), fill=(120, 128, 138))
    return img


def read_products_with_rows(xlsx_path):
    """Read every product row, returning dicts that carry their 1-indexed
    worksheet row so images can be matched by drawing anchor."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [c.value or f"Col{c.column}" for c in ws[1]]
    products = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        p = dict(zip(headers, values))
        if p.get('ID') is not None:
            p['_row'] = r          # 1-indexed worksheet row
            products.append(p)
    return products


def build_row_image_map(xlsx_path):
    """Map 1-indexed worksheet row → (PIL image bytes, ext) using the drawing
    anchors. This is the authoritative image↔product mapping."""
    wb = openpyxl.load_workbook(xlsx_path)  # keep drawings (no data_only)
    ws = wb.active
    row_to_image = {}
    for img in getattr(ws, "_images", []):
        try:
            ws_row = img.anchor._from.row + 1   # anchor row is 0-indexed
        except Exception:
            continue
        try:
            data = img._data()
        except Exception:
            continue
        ext = (getattr(img, "format", None) or "png").lower()
        # first anchor for a row wins (rows have at most one image here)
        row_to_image.setdefault(ws_row, (data, ext))
    return row_to_image


# ── Box geometry (real software-box proportions, NOT a flat slab) ─────────
# Front face is portrait (W x H); D is a healthy depth so the box never reads
# as a paper-thin slab after the viewer normalises its size.
BOX_W = 1.5   # x  (cover width)
BOX_H = 2.1   # y  (cover height)
BOX_D = 0.5   # z  (spine depth)

# trimesh's glTF exporter flips V on write, and three.js' GLTFLoader samples
# with v=0 at the image BOTTOM. Verified with a marker render: the cover must
# be authored with bottom vertices at v=0 so it appears upright (not flipped).
FLIP_V = True


def _prep_cover(texture_img: Image.Image) -> Image.Image:
    """EXIF-correct, de-alpha and return an RGB cover image."""
    texture_img.load()

    # Palette / grey-with-alpha images: promote to RGBA first so the alpha
    # fill below runs (a plain .convert('RGB') would bake transparent pixels to
    # a stray palette colour — often black — around the product).
    if texture_img.mode in ('P', 'LA', 'PA') or (
        texture_img.mode == 'L' and 'transparency' in texture_img.info
    ):
        texture_img = texture_img.convert('RGBA')

    # EXIF orientation
    try:
        from PIL.ExifTags import ORIENTATION
        exif = texture_img._getexif()
        if exif:
            o = exif.get(ORIENTATION)
            if o == 3:   texture_img = texture_img.rotate(180, expand=True)
            elif o == 6: texture_img = texture_img.rotate(270, expand=True)
            elif o == 8: texture_img = texture_img.rotate(90, expand=True)
    except Exception:
        pass

    # Handle alpha → fill transparent with nearest opaque colour (no black edges)
    if texture_img.mode == 'RGBA':
        arr = np.array(texture_img)
        rgb = arr[:, :, :3].copy()
        alpha = arr[:, :, 3]
        opaque = alpha > 128
        transparent = ~opaque
        if transparent.any() and opaque.any():
            idx = ndimage.distance_transform_edt(
                transparent, return_distances=False, return_indices=True
            )
            for c in range(3):
                ch = rgb[:, :, c]
                ch[transparent] = ch[idx[0][transparent], idx[1][transparent]]
                rgb[:, :, c] = ch
        texture_img = Image.fromarray(rgb, 'RGB')
    elif texture_img.mode != 'RGB':
        texture_img = texture_img.convert('RGB')
    return texture_img


def _edge_color(img: Image.Image):
    """Median colour of the cover's border → used for the spine/sides and the
    letterbox so the whole box reads as one cohesive object."""
    a = np.asarray(img)
    if a.ndim != 3:
        return (238, 240, 244)
    b = max(2, min(a.shape[0], a.shape[1]) // 40)
    strip = np.concatenate([
        a[:b].reshape(-1, 3), a[-b:].reshape(-1, 3),
        a[:, :b].reshape(-1, 3), a[:, -b:].reshape(-1, 3),
    ])
    return tuple(int(x) for x in np.median(strip, axis=0))


def _build_atlas(cover: Image.Image, side_color, size: int = 2048):
    """Composite a single texture atlas:
        left column  (u 0..cover_u) = the product cover, aspect-fit + letterboxed
        right column (u cover_u..1) = a flat panel of `side_color`
    Returns (atlas_image, cover_u)."""
    atlas = Image.new('RGB', (size, size), side_color)
    cover_w = int(round(size * (BOX_W / BOX_H)))   # front-face aspect region
    region = Image.new('RGB', (cover_w, size), side_color)
    cw, ch = cover.size
    scale = min(cover_w / cw, size / ch)
    nw, nh = max(1, int(round(cw * scale))), max(1, int(round(ch * scale)))
    resized = cover.resize((nw, nh), Image.Resampling.LANCZOS)
    region.paste(resized, ((cover_w - nw) // 2, (size - nh) // 2))
    atlas.paste(region, (0, 0))
    return atlas, cover_w / size


def apply_texture(glb_path, texture_img: Image.Image, output_path):
    """Build a fresh, correctly-UV'd software box and texture it with the
    product cover.

    The base `box.glb` ships a *packed* UV unwrap: every one of the 6 faces is
    mapped to a different (and differently-rotated) sub-rectangle of the 0..1
    texture. Painting the whole cover across that atlas — as the old code did —
    made each face sample a different rotated/mirrored slice, so back/side
    faces showed reversed gibberish. We therefore ignore box.glb's UVs and
    author our own geometry: the cover goes on the FRONT and BACK faces
    upright and un-mirrored; the spine/top/bottom get a flat brand colour
    sampled from the cover's edge. `glb_path` is kept for signature compat.
    """
    cover = _prep_cover(texture_img)
    side = _edge_color(cover)
    atlas, cu = _build_atlas(cover, side)
    atlas.save(os.path.join(os.path.dirname(output_path), 'texture.png'))

    hx, hy, hz = BOX_W / 2.0, BOX_H / 2.0, BOX_D / 2.0

    V, F, UV = [], [], []

    def add_face(quad, uvs):
        # quad given CCW as seen from OUTSIDE → outward normals, front-facing.
        i = len(V)
        V.extend(quad)
        UV.extend(uvs)
        F.append([i, i + 1, i + 2])
        F.append([i, i + 2, i + 3])

    # Cover UVs for [BL, BR, TR, TL] with glTF top-left origin.
    v_bottom, v_top = (0.0, 1.0) if FLIP_V else (1.0, 0.0)
    cover_uv = [(0.0, v_bottom), (cu, v_bottom), (cu, v_top), (0.0, v_top)]
    # A single point inside the flat side panel → uniform colour, no distortion.
    sp = (cu + (1.0 - cu) / 2.0, 0.5)
    side_uv = [sp, sp, sp, sp]

    # FRONT (+Z): cover, upright.  Quad order BL, BR, TR, TL.
    add_face([(-hx, -hy,  hz), ( hx, -hy,  hz), ( hx,  hy,  hz), (-hx,  hy,  hz)], cover_uv)
    # BACK (-Z): cover, upright & un-mirrored when viewed from behind.
    add_face([( hx, -hy, -hz), (-hx, -hy, -hz), (-hx,  hy, -hz), ( hx,  hy, -hz)], cover_uv)
    # RIGHT (+X): flat side panel.
    add_face([( hx, -hy,  hz), ( hx, -hy, -hz), ( hx,  hy, -hz), ( hx,  hy,  hz)], side_uv)
    # LEFT (-X): flat side panel.
    add_face([(-hx, -hy, -hz), (-hx, -hy,  hz), (-hx,  hy,  hz), (-hx,  hy, -hz)], side_uv)
    # TOP (+Y): flat side panel.
    add_face([(-hx,  hy,  hz), ( hx,  hy,  hz), ( hx,  hy, -hz), (-hx,  hy, -hz)], side_uv)
    # BOTTOM (-Y): flat side panel.
    add_face([(-hx, -hy, -hz), ( hx, -hy, -hz), ( hx, -hy,  hz), (-hx, -hy,  hz)], side_uv)

    mesh = trimesh.Trimesh(
        vertices=np.array(V, dtype=np.float64),
        faces=np.array(F, dtype=np.int64),
        process=False,
    )
    material = trimesh.visual.material.PBRMaterial(
        baseColorTexture=atlas, metallicFactor=0.0, roughnessFactor=0.85,
    )
    mesh.visual = trimesh.visual.TextureVisuals(
        uv=np.array(UV, dtype=np.float64), material=material,
    )
    mesh.export(str(output_path))


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    if not EXCEL_PATH.exists():
        sys.exit(f"ERROR: {EXCEL_PATH} not found")
    if not BASE_GLB.exists():
        sys.exit(f"ERROR: {BASE_GLB} not found")

    print("=" * 70)
    print("  DSM-3D  ·  Batch Product Processor  (anchor-based mapping)")
    print("=" * 70)

    products = read_products_with_rows(EXCEL_PATH)
    print(f"\n📦 Found {len(products)} products in {EXCEL_PATH.name}")
    print("🎨 Designing a uniform DSM cover for every product (data-driven).")

    MODELS_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_MODELS.mkdir(parents=True, exist_ok=True)

    # Preserve any DSM-Original (id 90000+) entries already in the manifest so a
    # catalog regen doesn't drop them (gen_own_boxes.py owns those rows).
    preserved = []
    if MANIFEST_PATH.exists():
        try:
            for m in json.load(open(MANIFEST_PATH)):
                try:
                    if int(m.get("id", 0)) >= 90000:
                        preserved.append(m)
                except (TypeError, ValueError):
                    pass
        except Exception:
            pass

    manifest = list(preserved)
    total = len(products)
    success = errors = 0

    for idx, product in enumerate(products):
        pid  = product.get('ID', 'unknown')
        name = str(product.get('Name') or f'product_{pid}')
        sku  = product.get('SKU')
        row  = product.get('_row')
        safe = sanitize(name)
        folder_name = f"{pid}_{safe}"
        folder_path = MODELS_DIR / folder_name
        output_glb  = folder_path / GLB_FILENAME
        flat_glb    = PUBLIC_MODELS / f"{pid}.glb"

        def record(status):
            manifest.append({
                "id":     pid,
                "name":   name,
                "sku":    sku,
                "file":   f"{pid}.glb" if status == "ok" else None,  # flat, site-loaded
                "folder": folder_name,
                "glb":    GLB_FILENAME if status == "ok" else None,  # api.py compat
                "status": status,
            })

        print(f"\n[{idx+1}/{total}] ID={pid}  row={row}  {name[:50]}")

        folder_path.mkdir(parents=True, exist_ok=True)
        try:
            params = derive_cover_params(product)
            print(f"   brand={params['bottom_title']!r} glyph={params['glyph']!r}"
                  f" badge={params['badge']!r}")
            cover = make_cover(**params)
            cover.save(folder_path / "cover.png")
            apply_texture(BASE_GLB, cover, output_glb)
            shutil.copy2(output_glb, flat_glb)   # publish flat copy for the site
            print(f"   ✅ → {output_glb.name}  +  public/models/{pid}.glb")
            record("ok")
            success += 1
        except Exception as e:
            print(f"   ❌ Error: {e}")
            traceback.print_exc()
            record(f"error: {e}")
            errors += 1

    with open(MANIFEST_PATH, 'w') as f:
        json.dump(manifest, f, indent=2, default=str)
    print(f"\n📝 Manifest written to {MANIFEST_PATH}")

    print("\n" + "=" * 70)
    print(f"  DONE  ·  ✅ {success} built  ❌ {errors} errors  "
          f"(+{len(preserved)} DSM-Original entries preserved)")
    print(f"  Flat models: {PUBLIC_MODELS}/")
    print(f"  Folder models: {MODELS_DIR}/")
    print("=" * 70)


if __name__ == "__main__":
    main()
