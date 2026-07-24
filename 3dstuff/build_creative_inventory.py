#!/usr/bin/env python3
"""Build the DSM creative-box inventory from extracted source artwork.

Writes a UI-friendly JSON registry plus an auditable Markdown checklist.  The
WordPress export is authoritative for the complete product catalogue; the live
creative manifest is authoritative for what is currently rendered by the API.
"""
from __future__ import annotations

import json
import re
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE / "creative-boxes"
WORDPRESS = HERE / "products.xlsx"
LIVE = HERE / "live-creative-manifest.json"
OUT_JSON = HERE.parent / "src" / "data" / "creativeInventory.json"
OUT_MD = HERE / "CREATIVE_BOXES_CHECKLIST.md"


def normalized(value: str) -> str:
    value = value.lower()
    value = re.sub(r"\b(front|back|spine|mak|license|licensing)\b", " ", value)
    value = re.sub(r"^\s*\d+\s*", "", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def title_for(path: Path) -> str:
    value = re.sub(r"\bfront\b", "", path.stem, flags=re.I)
    value = re.sub(r"^\s*\d+\s*", "", value)
    return re.sub(r"\s+", " ", value).strip(" -_")


def component_paths(front: Path) -> dict[str, str | None]:
    prefix = re.match(r"(\d+)\b", front.name)
    candidates = list(front.parent.glob("*.png"))
    matching = [p for p in candidates if prefix and p.name.startswith(prefix.group(1))]
    back = next((p for p in matching if "back" in p.name.lower()), None)
    spine = next((p for p in matching if "spine" in p.name.lower()), None)
    return {
        "front": front.relative_to(ROOT).as_posix(),
        "back": back.relative_to(ROOT).as_posix() if back else None,
        "spine": spine.relative_to(ROOT).as_posix() if spine else None,
    }


def collection_for(path: Path) -> tuple[str, str]:
    relative = path.relative_to(ROOT)
    if relative.parts[0] == "new-collections":
        return relative.parts[1], relative.parts[1]
    if relative.parts[0] == "autodesk":
        return "Autodesk 2027 Boxes.zip", "Previously extracted Autodesk 2027"
    if relative.parts[0] == "microsoft":
        return "Boxes 1-10.zip / Microsoft Windows 11 Boxes.zip", "Previously extracted Microsoft"
    return "Previously extracted", "Previously extracted"


def score(a: str, b: str) -> float:
    a_norm, b_norm = normalized(a), normalized(b)
    seq = SequenceMatcher(None, a_norm, b_norm).ratio()
    a_tokens, b_tokens = set(a_norm.split()), set(b_norm.split())
    overlap = len(a_tokens & b_tokens) / max(1, len(a_tokens | b_tokens))
    return round((seq * 0.62) + (overlap * 0.38), 3)


def wordpress_products() -> list[dict]:
    sheet = load_workbook(WORDPRESS, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    fields = next(rows)
    records = []
    for row in rows:
        item = dict(zip(fields, row))
        if item.get("ID") and item.get("Name"):
            records.append({"id": int(item["ID"]), "name": str(item["Name"]), "category": str(item.get("Categories") or "")})
    return records


def main() -> None:
    live = json.loads(LIVE.read_text()) if LIVE.exists() else []
    live_by_source = {entry.get("source"): entry for entry in live}
    products = wordpress_products()
    entries = []
    for front in sorted(path for path in ROOT.rglob("*.png") if "front" in path.name.lower()):
        title = title_for(front)
        product = max(products, key=lambda item: score(title, item["name"]))
        product_score = score(title, product["name"])
        matched = product if product_score >= 0.58 else None
        zip_name, collection = collection_for(front)
        components = component_paths(front)
        live_item = live_by_source.get(components["front"])
        entries.append({
            "collection": collection,
            "zip_name": zip_name,
            "title": title,
            "source": components,
            "wordpress_product": matched,
            "match_score": product_score,
            "live": bool(live_item),
            "live_model_id": live_item.get("id") if live_item else None,
            "checklist": {
                "front": True,
                "back": bool(components["back"]),
                "spine": bool(components["spine"]),
                "wordpress_match": bool(matched),
                "live_model": bool(live_item),
            },
        })

    matched_wp_ids = {entry["wordpress_product"]["id"] for entry in entries if entry["wordpress_product"]}
    missing = [product for product in products if product["id"] not in matched_wp_ids]
    collections = {}
    for entry in entries:
        bucket = collections.setdefault(entry["collection"], {"zip_name": entry["zip_name"], "total": 0, "live": 0})
        bucket["total"] += 1
        bucket["live"] += int(entry["live"])

    payload = {
        "generated_from": "WordPress products.xlsx + VPS creative-manifest.json",
        "wordpress_catalogue_count": len(products),
        "creative_source_count": len(entries),
        "live_creative_count": sum(1 for entry in entries if entry["live"]),
        "collections": [{"name": name, **data} for name, data in sorted(collections.items())],
        "designs": entries,
        "missing_wordpress_creatives": missing,
    }
    OUT_JSON.write_text(json.dumps(payload, indent=2) + "\n")

    lines = [
        "# DSM Creative Box Checklist",
        "",
        "Generated from the complete 478-product WordPress export and the live VPS creative manifest.",
        "",
        f"- WordPress catalogue products: **{len(products)}**",
        f"- Creative source fronts: **{len(entries)}**",
        f"- Currently live creative GLBs: **{payload['live_creative_count']}**",
        f"- Catalogue products without a matched creative source: **{len(missing)}**",
        "",
        "## Collection status",
        "",
        "| Collection / ZIP | Source fronts | Live now |",
        "| --- | ---: | ---: |",
    ]
    for item in payload["collections"]:
        lines.append(f"| {item['zip_name']} | {item['total']} | {item['live']} |")
    lines.extend(["", "## Design checklist", ""])
    for entry in entries:
        c = entry["checklist"]
        product = entry["wordpress_product"] or {}
        state = "live" if c["live_model"] else "not yet imported"
        lines.extend([
            f"### {entry['title']}",
            f"- Collection: `{entry['zip_name']}`",
            f"- Files: [x] front · [{'x' if c['back'] else ' '}] back · [{'x' if c['spine'] else ' '}] spine",
            f"- WordPress: {'[x] ' + product.get('name', '') if c['wordpress_match'] else '[ ] no confident match'}",
            f"- Live GLB: [{'x' if c['live_model'] else ' '}] {state}{' (ID ' + str(entry['live_model_id']) + ')' if c['live_model'] else ''}",
            "",
        ])
    lines.extend(["## Missing creative designs", ""])
    for product in missing:
        lines.append(f"- [ ] `{product['id']}` — {product['name']} ({product['category']})")
    OUT_MD.write_text("\n".join(lines) + "\n")
    print(json.dumps({"sources": len(entries), "live": payload["live_creative_count"], "missing": len(missing)}, indent=2))


if __name__ == "__main__":
    main()
