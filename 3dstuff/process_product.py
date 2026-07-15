#!/usr/bin/env python
"""
Script to extract a product from products.xlsx, create a folder for it,
and apply the product image as a texture on box.glb
"""

import os
import sys
import shutil
import zipfile
import tempfile
from pathlib import Path

# Try to import required libraries
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import trimesh
    HAS_TRIMESH = True
except ImportError:
    HAS_TRIMESH = False

try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# If libraries are missing, provide instructions
if not HAS_OPENPYXL:
    print("ERROR: openpyxl is required to read Excel files.")
    print("Please install it:")
    print("  pip install --user openpyxl")
    sys.exit(1)

if not HAS_TRIMESH:
    print("ERROR: trimesh is required to modify GLB files.")
    print("Please install it:")
    print("  pip install --user trimesh")
    sys.exit(1)

if not HAS_PIL:
    print("ERROR: PIL/Pillow is required to process images.")
    print("Please install it:")
    print("  pip install --user pillow")
    sys.exit(1)


def read_products_excel(file_path):
    """Read products from Excel file and extract images"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else f"Column_{cell.column}")
    
    # Extract images from Excel file
    image_map = {}  # Map row number to image path
    try:
        # Excel files are zip archives, extract images
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            # Find image files
            image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/') and (f.endswith('.jpg') or f.endswith('.jpeg') or f.endswith('.png'))]
            
            # Extract images to temp directory
            temp_dir = tempfile.mkdtemp()
            for img_file in image_files:
                zip_ref.extract(img_file, temp_dir)
                image_map[img_file] = os.path.join(temp_dir, img_file)
    except Exception as e:
        print(f"Warning: Could not extract images from Excel: {e}")
        temp_dir = None
    
    # Read data rows
    products = []
    row_num = 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        product = {}
        for i, value in enumerate(row):
            if i < len(headers):
                product[headers[i]] = value
        
        # Try to find associated image for this row
        if temp_dir:
            product['_row_num'] = row_num
            product['_temp_dir'] = temp_dir
            product['_image_files'] = list(image_map.values()) if image_map else []
        
        # Only add non-empty products
        if any(v is not None for v in product.values()):
            products.append(product)
        row_num += 1
    
    print("Products found:")
    for i, p in enumerate(products[:5]):  # Show first 5
        print(f"  Product {i}: {dict((k, v) for k, v in p.items() if not k.startswith('_'))}")
    print(f"\nTotal products: {len(products)}")
    print(f"Columns: {headers}")
    print(f"Found {len(image_map)} embedded images")
    return products, headers, image_map


def extract_product(products, index=0):
    """Extract one product from the list"""
    if index >= len(products):
        index = 0
    product = products[index]
    print(f"\nExtracting product at index {index}:")
    for key, value in product.items():
        print(f"  {key}: {value}")
    return product


def create_product_folder(product, base_dir="."):
    """Create a folder for the product"""
    product_name = None
    for col in ['name', 'product_name', 'product', 'id', 'product_id', 'title', 'Name', 'Product Name', 'Product', 'ID', 'Product ID', 'Title']:
        if col in product and product[col] is not None:
            product_name = str(product[col]).strip()
            if product_name:
                break
    
    if not product_name:
        for key, value in product.items():
            if value is not None:
                product_name = str(value).strip()
                if product_name:
                    break
    
    if not product_name:
        product_name = "product_unknown"
    
    # Clean folder name
    folder_name = "".join(c for c in product_name if c.isalnum() or c in (' ', '-', '_')).strip()
    folder_name = folder_name.replace(' ', '_')
    folder_path = Path(base_dir) / folder_name
    folder_path.mkdir(exist_ok=True)
    print(f"\nCreated folder: {folder_path}")
    return folder_path


def get_image_path(product, image_map=None, row_index=0):
    """Get the image path from product data or embedded images"""
    for col in ['image', 'image_path', 'image_url', 'img', 'picture', 'photo', 'thumbnail', 
                'Image', 'Image Path', 'Image URL', 'Img', 'Picture', 'Photo', 'Thumbnail']:
        if col in product and product[col] is not None:
            img_path = str(product[col]).strip()
            if img_path:
                if os.path.exists(img_path):
                    return img_path
                rel_path = os.path.join('.', img_path)
                if os.path.exists(rel_path):
                    return rel_path
                filename = os.path.basename(img_path)
                if os.path.exists(filename):
                    return filename
    
    if image_map and len(image_map) > 0:
        image_files = list(image_map.values())
        if row_index < len(image_files):
            img_path = image_files[row_index]
            if os.path.exists(img_path):
                return img_path
        if image_files and os.path.exists(image_files[0]):
            return image_files[0]
    
    if '_image_files' in product and product['_image_files']:
        img_files = product['_image_files']
        if row_index < len(img_files):
            img_path = img_files[row_index]
            if os.path.exists(img_path):
                return img_path
        if img_files and os.path.exists(img_files[0]):
            return img_files[0]
    
    return None


def apply_texture_to_glb(glb_path, texture_path, output_path):
    """Apply texture to GLB file while PRESERVING the original UV mapping.
    
    The original box.glb has carefully crafted UVs:
    - Front face maps to most of the texture (UV ~[0.2, 0.8] x [0.02, 0.98])
    - Side/top/bottom faces map to thin strips within the texture
    - All faces have UVs within the texture bounds — NO black areas
    - The tilt in the rendered output comes from the image content itself
    
    We simply swap the texture image while keeping the original UVs intact.
    """
    import numpy as np
    
    print(f"\nLoading GLB file: {glb_path}")
    scene = trimesh.load(glb_path)
    
    # NOTE: No geometry rotation needed — the tilt/perspective is already 
    # baked into the source image and the original UV mapping handles it.
    
    # Load and process the texture image
    print(f"Loading texture: {texture_path}")
    texture_img = Image.open(texture_path)
    print(f"  Image size: {texture_img.size}, mode: {texture_img.mode}")
    
    # Handle EXIF orientation
    try:
        from PIL.ExifTags import ORIENTATION
        exif = texture_img._getexif()
        if exif is not None:
            orientation = exif.get(ORIENTATION)
            if orientation == 3:
                texture_img = texture_img.rotate(180, expand=True)
                print("  Applied EXIF rotation: 180°")
            elif orientation == 6:
                texture_img = texture_img.rotate(270, expand=True)
                print("  Applied EXIF rotation: 270°")
            elif orientation == 8:
                texture_img = texture_img.rotate(90, expand=True)
                print("  Applied EXIF rotation: 90°")
    except:
        pass
    
    # CRITICAL: Handle RGBA images with transparency properly
    # The source image has transparent areas (53% transparent!) around the product box.
    # If we just convert RGBA→RGB, transparent pixels become BLACK, causing black edges.
    # Instead, we fill transparent areas with the nearest opaque pixel color.
    if texture_img.mode == 'RGBA':
        print("  Image has alpha channel — filling transparent areas...")
        img_array = np.array(texture_img)
        rgb = img_array[:, :, :3].copy()
        alpha = img_array[:, :, 3]
        
        # Find opaque pixels mask
        opaque_mask = alpha > 128
        transparent_mask = ~opaque_mask
        transparent_count = transparent_mask.sum()
        total_pixels = alpha.size
        print(f"  Transparent pixels: {transparent_count}/{total_pixels} ({100*transparent_count/total_pixels:.1f}%)")
        
        if transparent_count > 0 and transparent_count < total_pixels:
            # Fill transparent pixels by expanding outward from opaque regions
            # This creates a natural "bleed" that prevents black edges
            from scipy import ndimage
            
            # For each color channel, fill transparent pixels with nearest opaque value
            for c in range(3):
                channel = rgb[:, :, c].astype(float)
                # Use distance transform to find nearest opaque pixel for each transparent pixel
                # Set transparent pixels to 0, then dilate from opaque regions
                filled = channel.copy()
                
                # Iteratively expand opaque regions into transparent areas
                iterations = max(rgb.shape[0], rgb.shape[1])  # enough to fill everything
                indices = ndimage.distance_transform_edt(transparent_mask, return_distances=False, return_indices=True)
                filled[transparent_mask] = channel[indices[0][transparent_mask], indices[1][transparent_mask]]
                rgb[:, :, c] = filled.astype(np.uint8)
            
            print("  Filled transparent areas with nearest opaque pixel colors")
        
        texture_img = Image.fromarray(rgb, 'RGB')
    elif texture_img.mode != 'RGB':
        texture_img = texture_img.convert('RGB')
    
    print(f"  Processed image size: {texture_img.size}")
    
    # Resize to power-of-2 for GPU compatibility
    max_dim = max(texture_img.width, texture_img.height)
    power_of_2 = 2 ** ((max_dim - 1).bit_length())
    power_of_2 = max(512, min(2048, power_of_2))
    
    # Stretch the image to fill the full square texture
    final_texture = texture_img.resize((power_of_2, power_of_2), Image.Resampling.LANCZOS)
    
    print(f"  Final texture size: {final_texture.size}")
    
    # Save the processed texture for debugging
    texture_output = os.path.join(os.path.dirname(output_path), 'texture.png')
    final_texture.save(texture_output)
    print(f"  Saved texture to: {texture_output}")
    
    # Apply texture to the mesh — KEEP ORIGINAL UVs
    if hasattr(scene, 'geometry'):
        for name, mesh in scene.geometry.items():
            if hasattr(mesh, 'visual'):
                # Verify original UVs are present
                if hasattr(mesh.visual, 'uv') and mesh.visual.uv is not None:
                    uv = mesh.visual.uv
                    print(f"\n  Mesh '{name}': {len(mesh.vertices)} verts, {len(mesh.faces)} faces")
                    print(f"  Original UV range: [{uv.min(axis=0)}] to [{uv.max(axis=0)}]")
                    print(f"  Keeping original UVs — they define the box wrapping and tilt")
                else:
                    print(f"\n  Warning: Mesh '{name}' has no UVs, generating basic UVs...")
                    # Fallback: generate simple box projection UVs
                    vertices = mesh.vertices
                    bounds = mesh.bounds
                    size = bounds[1] - bounds[0]
                    center = (bounds[0] + bounds[1]) / 2
                    
                    uv_coords = np.zeros((len(vertices), 2))
                    for i, v in enumerate(vertices):
                        rel = v - center
                        uv_coords[i] = [(rel[0] / size[0] + 0.5), 1.0 - (rel[1] / size[1] + 0.5)]
                    
                    mesh.visual = trimesh.visual.TextureVisuals(uv=uv_coords)
                
                # Create PBR material with the new texture
                try:
                    material = trimesh.visual.material.PBRMaterial(
                        baseColorTexture=final_texture,
                        metallicFactor=0.0,
                        roughnessFactor=1.0,
                    )
                    mesh.visual.material = material
                    print(f"  Applied PBR texture to mesh: {name}")
                except Exception as e:
                    print(f"  Warning: PBR failed ({e}), trying SimpleMaterial...")
                    try:
                        material = trimesh.visual.material.SimpleMaterial(image=final_texture)
                        mesh.visual.material = material
                        print(f"  Applied SimpleMaterial to mesh: {name}")
                    except Exception as e2:
                        print(f"  Error applying texture: {e2}")
    
    # Export
    print(f"\nExporting GLB to: {output_path}")
    scene.export(output_path)
    print("Done! Texture applied with original UV mapping preserved.")


def main():
    excel_path = "products.xlsx"
    glb_path = "box.glb"
    
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found!")
        return
    
    if not os.path.exists(glb_path):
        print(f"Error: {glb_path} not found!")
        return
    
    # Read products
    products, headers, image_map = read_products_excel(excel_path)
    
    if not products:
        print("Error: No products found in Excel file!")
        return
    
    # Extract product (change index to test different products)
    # 0 = Chaos Corona Renderer, 3 = Visual Studio 2026 Professional
    # 5 = Autodesk Construction Cloud, 6 = Chaos Enscape
    product_index = 0
    product = extract_product(products, index=product_index)
    
    # Create folder
    product_folder = create_product_folder(product)
    
    # Get image path
    image_path = get_image_path(product, image_map, product_index)
    if not image_path:
        print("\nWarning: Could not find image path in product data.")
        print("Available columns:", list(product.keys()))
        return
    
    if not os.path.exists(image_path):
        print(f"\nError: Image file not found: {image_path}")
        return
    
    # Copy image to product folder
    image_filename = os.path.basename(image_path)
    dest_image_path = product_folder / image_filename
    shutil.copy2(image_path, dest_image_path)
    print(f"Copied image to: {dest_image_path}")
    
    # Copy GLB to product folder
    dest_glb_path = product_folder / "box.glb"
    shutil.copy2(glb_path, dest_glb_path)
    print(f"Copied GLB to: {dest_glb_path}")
    
    # Apply texture — preserving original UVs (no rotation, no UV regeneration)
    output_glb_path = product_folder / "box_textured.glb"
    try:
        apply_texture_to_glb(dest_glb_path, dest_image_path, output_glb_path)
        print(f"\n✓ Product processed successfully!")
        print(f"  Folder: {product_folder}")
        print(f"  Textured GLB: {output_glb_path}")
    except Exception as e:
        print(f"\nError applying texture: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
